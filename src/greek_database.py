#####################
# greek_database.py #
#####################

from openpyxl import Workbook, load_workbook

# ------------------------------------------------------------------------------

def process_greek_db(filename_origin = "data/Greece/original-db-greece.xlsx",
                     filename_dest = "data/Greece/processed-db-greece.xlsx"):

    # Leemos el libro de excel
    wb = load_workbook(filename_origin)

    # Obtenemos la primera hoja del documento
    sheet = wb.worksheets[0]

    # Obtenemos el número de columnas que hay en total
    n_colum = sheet.max_column

    # Insertamos una columa al final para añadir la region
    sheet.insert_cols(n_colum+1)
    cell_country = sheet.cell(row = 2, column = n_colum+1)
    cell_country.value = "Country"

    # Insertamos la fila que contendrá las unidades
    sheet.insert_rows(3)

    correspondences_values = ["Food name Spanish", "Food name English",
    "Energy (kcal)", "Protein", "Total lipids/fat", "Saturated fatty acids",
    "Monounsaturated fatty acids",	"Polyunsaturated fatty acids", "Carbohydrates",
    "Dietary fibre", "Water", "Sodium", "Potassium", "Calcium",	"Magnesium",
    "Phosphorus", "Iron", "Zinc", "Copper"]

    units_values = [" ", " ", "kcal", "g", "g", "g", "g", "g", "g", "g", "g",
     "mg", "mg", "mg", "mg", "mg", "mg", "mg", "mg", " "]

    # La primera celda de la tabla es la (1,1), no la (0,0)
    for i in range(1, n_colum+1):
        cell_names = sheet.cell(row = 2, column = i)
        cell_names.value = correspondences_values[i-1]

        cell_units = sheet.cell(row = 3, column = i)
        cell_units.value = units_values[i-1]

    wb.save(filename_dest)

# ------------------------------------------------------------------------------

if __name__ == "__main__":

    process_greek_db()
